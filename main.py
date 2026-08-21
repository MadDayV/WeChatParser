import asyncio
import json
import time
import os
import re
from typing import Dict, Any, List, Optional, Tuple
import httpx
from pydantic import BaseModel, Field, field_validator, model_validator
from dotenv import load_dotenv
import sys
import argparse
from datetime import datetime
import logging

# Загружаем конфигурацию из .env
load_dotenv()

# ==================== СЧИТЫВАНИЕ CONFIG ИЗ .env ====================
ALBUM_ID = os.getenv("ALBUM_ID", "_ZZajIW0nut8N3cyxzlhAGRf9BcyL4ZIU")
SHOP_ID = os.getenv("SHOP_ID", "").strip()
TARGET_URL = f"https://www.szwego.com/album/personal/all"

AI_PROVIDER = os.getenv("AI_PROVIDER", "openrouter").lower()
TOTAL_TARGET_PRODUCTS = int(os.getenv("TOTAL_TARGET_PRODUCTS", 100))
SZWEGO_SAFE_PAUSE = float(os.getenv("SZWEGO_SAFE_PAUSE", 12.0))

# Считываем токен из .env
SZWEGO_TOKEN = os.getenv("SZWEGO_TOKEN")
# Файл-черновик для хранения поштучно собранных ссылок
RAW_DUMP_FILE = "raw_selected_products.json"

if not SZWEGO_TOKEN:
    raise ValueError("[CRITICAL-ERROR] Токен SZWEGO_TOKEN не найден в файле .env!")

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")

# Считываем модель из .env. По умолчанию — точная gemini-3.1-pro-preview в формате OpenRouter
OPENROUTER_MODEL = os.getenv("OPENROUTER_MODEL", "google/gemini-3.1-pro-preview")
# Модель для нативного провайдера Gemini (без OpenRouter)
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-3.1-pro-preview")
# Температура генерации: ниже = меньше выдумок (важно для артикулов)
AI_TEMPERATURE = float(os.getenv("AI_TEMPERATURE", 0.4))
# ===================================================================
# ИИ-Клиенты
openai_client = None
openrouter_client = None
gemini_client = None

# Путь к файлу промпта в папке скрипта
PROMPT_FILE = os.path.join(os.path.dirname(__file__), "ai_prompt.txt")

def load_system_prompt() -> str:
    """Безопасно загружает промпт из TXT файла с резервным фоллбеком"""
    if os.path.exists(PROMPT_FILE):
        try:
            with open(PROMPT_FILE, "r", encoding="utf-8") as f:
                content = f.read().strip()
                if content:
                    return content
        except Exception as e:
            print(f"[WARNING] Не удалось прочитать ai_prompt.txt: {e}. Применяю дефолтный промпт.")
    
    # Резервный хардкод на случай, если Павел удалит файл
    return (
        "Ты — эксперт по контенту для премиальных брендов одежды и аксессуаров. "
        "Переведи китайский текст поставщика в карточку на русском. "
        "В 'title_ru_short' держи формат: [Вид товара] - [Бренд] - [Модель]."
    )

# Инициализация выбранного ИИ
if AI_PROVIDER == "openai":
    from openai import AsyncOpenAI
    if not OPENAI_API_KEY:
        raise ValueError("[ERROR] Выбран OpenAI, но OPENAI_API_KEY не найден в .env!")
    openai_client = AsyncOpenAI(api_key=OPENAI_API_KEY)
    print("[INIT] Режим: Боевой OpenAI (GPT-4o-mini)")

elif AI_PROVIDER == "openrouter":
    from openai import AsyncOpenAI
    if not OPENROUTER_API_KEY:
        raise ValueError("[ERROR] Выбран OpenRouter, но OPENROUTER_API_KEY не найден в .env!")
    # Подменяем базовый URL для шлюза OpenRouter
    openrouter_client = AsyncOpenAI(
        base_url="https://openrouter.ai",
        api_key=OPENROUTER_API_KEY,
        default_headers={
            "HTTP-Referer": "https://github.com",  # Для аналитики OpenRouter
            "X-Title": "Szwego AI Parser"
        }
    )
    print("[INIT] Режим: Боевой OpenRouter (Шлюз ИИ подключен)")

elif AI_PROVIDER == "gemini":
    from google import genai
    if not GEMINI_API_KEY:
        raise ValueError("[ERROR] Выбран Gemini, но GEMINI_API_KEY не найден в .env!")
    gemini_client = genai.Client(api_key=GEMINI_API_KEY)
    from google.genai import types
    print(f"[INIT] Режим: Боевой Google Gemini ({GEMINI_MODEL})")

else:
    print("[INIT] Режим: Локальный автономный MOCK (ИИ отключен, лимиты не тратятся)")

class ProductCard(BaseModel):
    title_ru_short: str = Field(description="Короткое название товара на русском языке")
    title_ru: str = Field(description="SEO-оптимизированное название товара на русском языке")
    brand: str = Field(description="Бренд товара (название производителя или торговой марки)")
    sku: str = Field(description="Артикул или уникальный код модели")
    description_ru: str = Field(description="Описание товара для покупателя")
    category: str = Field(description="Категория товара")
    tags: List[str] = Field(description="Список ключевых слов")

    @model_validator(mode="before")
    @classmethod
    def normalize_ai_payload(cls, data: Any) -> Any:
        if not isinstance(data, dict):
            return data

        if "product_card" in data and isinstance(data["product_card"], dict):
            data = data["product_card"]

        aliases = {
            "title": "title_ru",
            "description": "description_ru",
            "category_ru": "category",
            "short_title": "title_ru_short",
        }
        for old_key, new_key in aliases.items():
            if old_key in data and new_key not in data:
                data[new_key] = data[old_key]

        title = data.get("title_ru") or data.get("title_ru_short") or data.get("title") or "Товар"
        data.setdefault("title_ru_short", title)
        data.setdefault("title_ru", title)
        data.setdefault("description_ru", data.get("description") or "")
        data.setdefault("brand", "Премиальный бренд")
        data.setdefault("sku", "")
        data.setdefault("category", "")
        data.setdefault("tags", [])

        return data

    @field_validator("tags", mode="before")
    @classmethod
    def coerce_tags(cls, value: Any) -> List[str]:
        if value is None:
            return []
        if isinstance(value, str):
            return [tag.strip() for tag in re.split(r"[,;|]", value) if tag.strip()]
        if isinstance(value, list):
            return [str(tag).strip() for tag in value if str(tag).strip()]
        return []

    @field_validator("sku", "brand", mode="before")
    @classmethod
    def coerce_required_strings(cls, value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        if text.upper() == "N/A":
            return ""
        return text

    @field_validator("category", mode="before")
    @classmethod
    def coerce_category(cls, value: Any) -> str:
        if value is None:
            return ""
        return normalize_category_path(str(value).strip())

    @model_validator(mode="after")
    def fill_empty_strings(self) -> "ProductCard":
        # sku: пустое поле, если артикул не найден (не N/A)
        if self.sku and self.sku.upper() == "N/A":
            object.__setattr__(self, "sku", "")
        if not self.brand:
            object.__setattr__(self, "brand", "Неизвестный производитель")
        if not self.tags:
            fallback_tags = [
                tag for tag in (self.brand, self.category)
                if tag and tag not in ("Неизвестный производитель",)
            ]
            object.__setattr__(self, "tags", fallback_tags or ["lux"])
        return self

HEADERS = {
    "accept": "application/json, text/javascript, */*; q=0.01",
        "accept-language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
        "content-type": "application/x-www-form-urlencoded; charset=UTF-8",
        "origin": "https://www.szwego.com",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36",
        "x-requested-with": "XMLHttpRequest",
        # Специфичные заголовки платформы Szwego
        "wego-channel": "net",
        "wego-staging": "1",
        "x-wg-language": "zh"
}

COOKIES = {
    'token': SZWEGO_TOKEN
}

def extract_json_object(text: str) -> str:
    """Вырезает первый JSON-объект из ответа ИИ (без markdown)."""
    cleaned = text.replace("```json", "").replace("```", "").strip()
    start = cleaned.find("{")
    end = cleaned.rfind("}")
    if start == -1 or end == -1 or end <= start:
        raise ValueError("В ответе ИИ не найден JSON-объект {...}")
    return cleaned[start : end + 1]


def repair_json_control_chars(text: str) -> str:
    """
    Чинит типичные поломки JSON от LLM:
    - сырые переносы строк внутри кавычек (Invalid control character)
    - битые escape-последовательности (Invalid \\escape)
    """
    result: List[str] = []
    in_string = False
    i = 0
    n = len(text)

    while i < n:
        ch = text[i]

        if not in_string:
            if ch == '"':
                in_string = True
            result.append(ch)
            i += 1
            continue

        # внутри строки
        if ch == "\\":
            nxt = text[i + 1] if i + 1 < n else ""
            if nxt in '"\\/bfnrt':
                result.append(ch)
                result.append(nxt)
                i += 2
                continue
            if nxt == "u" and i + 5 < n and all(
                c in "0123456789abcdefABCDEF" for c in text[i + 2 : i + 6]
            ):
                result.append(text[i : i + 6])
                i += 6
                continue
            # Битый escape: экранируем сам слэш
            result.append("\\\\")
            i += 1
            continue

        if ch == '"':
            in_string = False
            result.append(ch)
            i += 1
            continue

        if ch == "\n":
            result.append("\\n")
            i += 1
            continue
        if ch == "\r":
            result.append("\\r")
            i += 1
            continue
        if ch == "\t":
            result.append("\\t")
            i += 1
            continue
        if ord(ch) < 32:
            result.append(f"\\u{ord(ch):04x}")
            i += 1
            continue

        result.append(ch)
        i += 1

    return "".join(result)


def parse_ai_json_payload(raw_text: str) -> dict:
    """Парсит ответ ИИ в dict: сначала напрямую, при ошибке — с починкой."""
    candidate = extract_json_object(raw_text)

    try:
        parsed = json.loads(candidate)
    except json.JSONDecodeError:
        repaired = repair_json_control_chars(candidate)
        parsed = json.loads(repaired)

    if isinstance(parsed, dict) and "product_card" in parsed and isinstance(parsed["product_card"], dict):
        return parsed["product_card"]
    if isinstance(parsed, dict):
        return parsed
    raise ValueError("Ответ ИИ не является JSON-объектом")


async def generate_ai_card(raw_text: str, max_retries: int = 3) -> Optional[ProductCard]:
    if not raw_text or not raw_text.strip():
        return None
        
    # Динамически подгружаем актуальный промпт перед каждым запросом к ИИ
    system_prompt = load_system_prompt()
    user_prompt = (
        "Ниже данные ОДНОГО конкретного товара с SZWEGO.\n"
        "Пиши карточку только про него. Примеры из инструкции копировать нельзя.\n"
        "Если по этим данным нельзя уверенно написать описание — description_ru должен быть пустой строкой \"\".\n\n"
        f"Данные:\n{raw_text}"
    )
    
    for attempt in range(1, max_retries + 1):
        try:
            if AI_PROVIDER == "mock":
                await asyncio.sleep(0.05)
                sku_match = re.search(r'[A-Za-z0-9\-_]{4,15}', raw_text)
                extracted_sku = sku_match.group(0) if sku_match else "Brand-Lux"
                return ProductCard(
                    title_ru_short=f"Премиальная модель {extracted_sku}",
                    title_ru=f"Премиальная модель {extracted_sku}",
                    brand="Премиальный бренд",
                    sku=extracted_sku,
                    description_ru=f"Локальный тест. Сырой текст: {raw_text.strip()[:40]}...",
                    category="Одежда и Аксессуары",
                    tags=["mock", "test", "lux"]
                )
    
            elif AI_PROVIDER == "gemini":
                loop = asyncio.get_event_loop()
                
                # Явный вызов синхронной функции в пуле потоков без ломающих lambda-оберток
                response = await loop.run_in_executor(
                    None,
                    lambda: gemini_client.models.generate_content(
                        model=GEMINI_MODEL,
                        contents=user_prompt,
                        config=types.GenerateContentConfig(
                            system_instruction=system_prompt,
                            response_mime_type="application/json",
                            response_schema=ProductCard,
                            temperature=AI_TEMPERATURE
                        )
                    )
                )
                # Важно: берем текст ответа и валидируем через Pydantic-модель
                return ProductCard.model_validate_json(response.text)
                
            elif AI_PROVIDER == "openai":
                response = await openai_client.beta.chat.completions.parse(
                    model="gpt-4o-mini",
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_prompt}
                    ],
                    response_format=ProductCard,
                    temperature=AI_TEMPERATURE,
                    timeout=30.0
                )
                # ИСПРАВЛЕНО: добавлен индекс [0] для выбора первого варианта ответа
                return response.choices[0].message.parsed
            
            elif AI_PROVIDER == "openrouter":
                headers = {
                    "Authorization": f"Bearer {OPENROUTER_API_KEY}",
                    "Content-Type": "application/json",
                    "HTTP-Referer": "https://github.com",
                    "X-Title": "Szwego AI Parser"
                }

                # Пример — ТОЛЬКО ключи JSON. Значения пустые: их нельзя копировать в карточку.
                json_example = (
                    "{\n"
                    '  "title_ru_short": "",\n'
                    '  "title_ru": "",\n'
                    '  "brand": "",\n'
                    '  "sku": "",\n'
                    '  "description_ru": "",\n'
                    '  "category": "",\n'
                    '  "tags": []\n'
                    "}"
                )

                mcp_system_content = (
                    f"{system_prompt}\n"
                    f"Выдай ответ СТРОГО в формате JSON, соответствующем этой структуре ключей:\n{json_example}\n"
                    f"КРИТИЧЕСКИЕ ПРАВИЛА:\n"
                    f"1. Не используй markdown разметку ```json и ```. Начни ответ сразу с {{ и закончи }}.\n"
                    f"2. Все ключи в JSON должны быть плоскими (без вложений).\n"
                    f"3. Поле tags — только JSON-массив строк, не строка через запятую.\n"
                    f"4. Поля sku и category всегда строки.\n"
                    f"5. АРТИКУЛ (sku): если нет подтверждённого оригинального артикула — пиши строго пустую строку \"\". "
                    f"ЗАПРЕЩЕНО писать N/A, выдумывать код или копировать его из примера.\n"
                    f"6. КАТЕГОРИЯ (category): ТОЛЬКО полный путь из дерева категорий промпта, "
                    f"разделитель '>' БЕЗ пробелов вокруг, "
                    f"например \"Женский>Аксессуары>Сумки и рюкзаки>Через плечо\". "
                    f"ЗАПРЕЩЕНО короткие названия вроде \"Сумки\", \"Обувь\", \"Одежда и Аксессуары\".\n"
                    f"7. В description_ru переносы строк пиши ТОЛЬКО как экранированные \\n внутри JSON-строки. "
                    f"Нельзя вставлять реальные переносы строк внутри кавычек JSON.\n"
                    f"8. ОПИСАНИЕ (description_ru): пиши ТОЛЬКО про товар из блока «Данные». "
                    f"Если данных мало и нельзя уверенно описать ИМЕННО эту вещь — description_ru строго \"\". "
                    f"ЗАПРЕЩЕНО копировать примеры из инструкции (сумки Louis Vuitton Speedy, часы Alhambra и любые другие примеры). "
                    f"Пустое поле лучше выдумки.\n"
                    f"9. title_ru_short / title_ru / brand: не подставляй «Сумка Louis Vuitton», если в данных не сумка "
                    f"или бренд не указан. Нет уверенности — пустые строки."
                )

                payload = {
                    "model": OPENROUTER_MODEL,
                    "messages": [
                        {"role": "system", "content": mcp_system_content},
                        {"role": "user", "content": user_prompt}
                    ],
                    "temperature": AI_TEMPERATURE,
                    "response_format": {"type": "json_object"},
                }
                
                async with httpx.AsyncClient(timeout=90.0) as client:
                    res = await client.post("https://openrouter.ai/api/v1/chat/completions", headers=headers, json=payload)
                    
                    # Если OpenRouter вернул ошибку, мы выведем её ТЕКСТ в консоль, а не упадем
                    if res.status_code != 200:
                        raise ValueError(f"OpenRouter Error {res.status_code}: {res.text}")
                        
                    data = res.json()
                    raw_json_text = data["choices"][0]["message"]["content"]
                    parsed_dict = parse_ai_json_payload(raw_json_text)
                    return ProductCard.model_validate(parsed_dict)
                
        except Exception as e:
            print(f"[Попытка {attempt}/{max_retries}] Ошибка вызова ИИ: {e}")
            if attempt < max_retries:
                await asyncio.sleep(2)  # Небольшая пауза перед повтором при ошибке сети/лимитов
            else:
                print(f"❌ Не удалось обработать товар после {max_retries} попыток.")
                
    if not raw_text or not raw_text.strip():
        return None

async def fetch_catalog_page(target_timestamp: Optional[int] = None) -> Dict[str, Any]:
    current_timestamp = target_timestamp if target_timestamp is not None else int(time.time() * 1000)
    params = {
        'albumId': str(ALBUM_ID), 'searchValue': '', 'searchImg': '', 'startDate': '', 'endDate': '',
        'sourceId': '', 'slipType': '1', 'timestamp': str(current_timestamp), 'requestDataType': '', 'transLang': 'en'
    }
    data = {'tagList': '[]'}
    
    async with httpx.AsyncClient(headers=HEADERS, cookies=COOKIES, timeout=30.0) as client:
        try:
            response = await client.post(TARGET_URL, params=params, data=data, follow_redirects=True)
            if response.status_code == 200:
                return response.json()
            return {}
        except Exception:
            return {}

async def process_and_generate_catalog(goods_list: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    final_catalog = []
    total_items = len(goods_list)
    print(f"\n[STAGE-2] Конвейер ИИ для {total_items} товаров (Провайдер: {AI_PROVIDER.upper()})...")
    
    for index, item in enumerate(goods_list, 1):
        goods_id = item.get("id") or item.get("goods_id")
        raw_text = item.get("content", "") or item.get("title", "")
        images = item.get("imgs_Src", []) or item.get("imgs", [])
        
        print(f"[AI] Обработка {index}/{total_items} (ID: {goods_id})...")
        ai_card = await generate_ai_card(raw_text)
        
        if ai_card:
            product_data = ai_card.model_dump()
            product_data["original_images"] = images
            product_data["szwego_id"] = goods_id
            final_catalog.append(product_data)
            print(f" -> [OK] {product_data['title_ru']} (SKU: {product_data['sku']})")
        else:
            print(f" -> [SKIP] Ошибка ИИ на товаре ID: {goods_id}")
    return final_catalog

async def main(target_limit: int = None):
    all_raw_goods = []
    next_timestamp_cursor = None
    
    # Подставляем лимит: если передан из батника — берем его, иначе — дефолт из .env
    limit = target_limit if target_limit is not None else TOTAL_TARGET_PRODUCTS
    
    print(f"[START] Парсер запущен из .env настроек.")
    print(f"Цель: {limit} товаров из альбома {ALBUM_ID}. Пауза: {SZWEGO_SAFE_PAUSE} сек.")
    
    # Заменяем старый TOTAL_TARGET_PRODUCTS на нашу новую переменную limit
    while len(all_raw_goods) < limit:
        raw_data = await fetch_catalog_page(target_timestamp=next_timestamp_cursor)
        if not raw_data:
            break
            
        result_data = raw_data.get("result", {})
        page_items = result_data.get("items", []) if isinstance(result_data, dict) else []
        
        if not page_items:
            print("[INFO] Достигнут конец каталога.")
            break
            
        print(f"[SZWEGO] Собрано {len(page_items)} позиций со страницы.")
        all_raw_goods.extend(page_items)
        
        next_timestamp_cursor = page_items[-1].get("update_time")
        if not next_timestamp_cursor:
            break
            
        if len(all_raw_goods) < limit:
            print(f"[SAFE-MODE] Защитная пауза {SZWEGO_SAFE_PAUSE} секунд...")
            await asyncio.sleep(SZWEGO_SAFE_PAUSE)
            
    all_raw_goods = all_raw_goods[:limit]
    print(f"\n[STAGE-1 COMPLETED] Сбор окончен. Позиций в пуле: {len(all_raw_goods)}")
    
    if not all_raw_goods:
        return
        
    final_catalog = await process_and_export_table(all_raw_goods)
    
    if final_catalog:
        with open("final_products.json", "w", encoding="utf-8") as f:
            json.dump(final_catalog, f, indent=2, ensure_ascii=False)

def extract_item_id(url: str) -> Optional[str]:
    """Вытаскивает id товара (последний сегмент после theme_detail/ или goods_detail/)."""
    match = re.search(r"(?:theme_detail|goods_detail)/[^/]+/([^/?#]+)", url)
    return match.group(1) if match else None

def extract_shop_id_from_url(url: str) -> Optional[str]:
    """shop_id может быть в query (?shop_id=...) или в поддомене (a123....szwego.com)."""
    shop_match = re.search(r"[?&]shop_id=([^&]+)", url)
    if shop_match:
        return shop_match.group(1)

    subdomain_match = re.search(r"https?://([a-zA-Z0-9_-]+)\.szwego\.com", url)
    if subdomain_match:
        subdomain = subdomain_match.group(1)
        if subdomain not in ("www", "m", "api", "szwego"):
            return subdomain

    return None


def normalize_shop_id(value: str) -> str:
    value = value.strip()
    if re.match(r"^A\d+$", value):
        return f"a{value[1:]}"
    return value


def is_numeric_shop_id(value: str) -> bool:
    return bool(re.match(r"^[Aa]\d{10,}$", value))


def resolve_shop_id_candidates(url: str) -> List[Optional[str]]:
    """Список shop_id для перебора: из URL, .env, либо запрос без shopId (None)."""
    candidates: List[Optional[str]] = []

    url_shop_id = extract_shop_id_from_url(url)
    if url_shop_id:
        candidates.append(normalize_shop_id(url_shop_id))
    else:
        if SHOP_ID:
            candidates.append(normalize_shop_id(SHOP_ID))
        if is_numeric_shop_id(ALBUM_ID):
            normalized_album_shop = normalize_shop_id(ALBUM_ID)
            if normalized_album_shop not in candidates:
                candidates.append(normalized_album_shop)
        candidates.append(None)

    seen = set()
    unique_candidates: List[Optional[str]] = []
    for candidate in candidates:
        key = candidate or ""
        if key in seen:
            continue
        seen.add(key)
        unique_candidates.append(candidate)

    return unique_candidates


async def _request_commodity_view(
    target_album_id: str,
    item_id: str,
    shop_id: Optional[str] = None,
) -> Dict[str, Any]:
    api_url = (
        f"https://szwego.com/commodity/view"
        f"?targetAlbumId={target_album_id}"
        f"&itemId={item_id}"
        f"&transLang=en"
    )
    if shop_id:
        api_url += f"&shopId={shop_id}"

    local_headers = {
        "accept": "application/json, text/plain, */*",
        "accept-language": "ru,ru-RU;q=0.9,en-US;q=0.8,en;q=0.7",
        "bundle_id": "",
        "cache-control": "no-cache",
        "pragma": "no-cache",
        "priority": "u=1, i",
        "sec-ch-ua": '"Chromium";v="148", "Google Chrome";v="148", "Not/A)Brand";v="99"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-origin",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36",
        "wego-albumid": "",
        "wego-channel": "net",
        "wego-staging": "0",
        "wego-uuid": "",
        "wego-version": "",
        "x-wg-language": "zh"
    }

    local_cookies = {
        "token": str(SZWEGO_TOKEN),
        "googtrans": "/en/ru"
    }

    async with httpx.AsyncClient(headers=local_headers, cookies=local_cookies, timeout=15.0) as client:
        response = await client.get(api_url, follow_redirects=False)

        if response.status_code != 200:
            print(f"⚠️ Неожиданный ответ сервера. Код: {response.status_code}. Текст: {response.text[:200]}")
            return {}

        res_json = response.json()
        if res_json.get("errcode") != 0:
            shop_hint = shop_id or "(без shopId)"
            print(f"⚠️ Szwego API [{shop_hint}]: {res_json.get('errmsg')}")
            return {}

        return res_json.get("result", res_json)


def parse_szwego_url(url: str) -> Dict[str, Optional[str]]:
    """Разбирает ссылку Szwego: товар (theme_detail) или магазин (shop_detail)."""
    shop_id = extract_shop_id_from_url(url)

    product_match = re.search(r"(?:theme_detail|goods_detail)/([^/]+)/([^/?#]+)", url)
    if product_match:
        return {
            "link_type": "product",
            "target_album_id": product_match.group(1),
            "item_id": product_match.group(2),
            "shop_id": shop_id,
        }

    shop_match = re.search(r"shop_detail/([^/?#]+)", url)
    if shop_match:
        return {
            "link_type": "shop",
            "target_album_id": shop_match.group(1),
            "item_id": None,
            "shop_id": shop_id,
        }

    return {
        "link_type": "unknown",
        "target_album_id": None,
        "item_id": None,
        "shop_id": shop_id,
    }


def print_shop_link_hint(album_id: str) -> None:
    print("❌ Это ссылка на МАГАЗИН (shop_detail), а не на конкретный товар.")
    print("   Режим «поштучного сбора» работает только со ссылкой на товар.")
    print()
    print("   Как получить правильную ссылку:")
    print("   1. Откройте магазин в браузере")
    print("   2. Кликните на нужный товар")
    print("   3. Скопируйте ссылку — в ней должен быть theme_detail и ДВА ID:")
    print(f"      ...#/theme_detail/{album_id}/XXXXX_товара")
    print()
    print("   Либо используйте «Массовый парсинг каталога» и укажите в .env:")
    print(f'      ALBUM_ID="{album_id}"')


async def fetch_single_item_raw_url(original_url: str) -> Dict[str, Any]:
    """
    Вытаскивает параметры из браузерной ссылки и делает прямой запрос 
    к эндпоинту /commodity/view на основном домене www.szwego.com.
    """
    parsed = parse_szwego_url(original_url)

    if parsed["link_type"] == "shop":
        print_shop_link_hint(parsed["target_album_id"] or "")
        return {}

    if parsed["link_type"] != "product":
        print("❌ Ошибка: Не удалось распознать ссылку Szwego!")
        print("   Ожидается формат: ...#/theme_detail/АЛЬБОМ/ТОВАР")
        return {}

    target_album_id = parsed["target_album_id"]
    item_id = parsed["item_id"]
    shop_candidates = resolve_shop_id_candidates(original_url)

    if extract_shop_id_from_url(original_url) is None:
        print(f"ℹ️ shop_id в ссылке не найден — пробуем {len(shop_candidates)} вариант(ов) запроса...")

    try:
        for shop_id in shop_candidates:
            item_data = await _request_commodity_view(target_album_id, item_id, shop_id)
            if item_data:
                if shop_id:
                    print(f"✅ Товар получен (shopId={shop_id})")
                else:
                    print("✅ Товар получен (запрос без shopId)")
                return item_data

        print("❌ Не удалось получить данные товара ни одним из способов.")
        if not extract_shop_id_from_url(original_url) and not SHOP_ID and not is_numeric_shop_id(ALBUM_ID):
            print("   Подсказка: добавьте в .env SHOP_ID=a201903291406004270013266")
            print("   (ID магазина из поддомена ссылки или из DevTools → Network).")
        return {}
    except Exception as e:
        print(f"\n❌ Ошибка сети при поштучном запросе: {e}")
        return {}

def load_raw_dump() -> list:
    """Читает черновик поштучного сбора; при ошибке/отсутствии — пустой список."""
    if not os.path.exists(RAW_DUMP_FILE):
        return []
    try:
        with open(RAW_DUMP_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def save_raw_dump(data: list) -> None:
    """Перезаписывает черновик поштучного сбора."""
    with open(RAW_DUMP_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)


def _normalize_img_list(img_urls) -> List[str]:
    """Приводит imgsSrc/imgs к списку строковых URL."""
    if not img_urls:
        return []
    if not isinstance(img_urls, list):
        return [str(img_urls)]
    if img_urls and isinstance(img_urls[0], dict):
        return [
            str(img.get("url") or img.get("thumb") or "").strip()
            for img in img_urls
            if img.get("url") or img.get("thumb")
        ]
    return [str(u).strip() for u in img_urls if str(u).strip()]


def _commodity_block(item: dict) -> dict:
    """Достаёт блок commodity из сырой записи (плоской или вложенной)."""
    if not isinstance(item, dict):
        return {}
    commodity = item.get("commodity")
    if isinstance(commodity, dict) and commodity:
        return commodity
    result = item.get("result") if isinstance(item.get("result"), dict) else {}
    nested = result.get("commodity") if isinstance(result.get("commodity"), dict) else {}
    return nested if isinstance(nested, dict) else {}


def extract_item_text(item: dict) -> Tuple[str, str]:
    """Достаёт (title, description) из сырого товара."""
    if not isinstance(item, dict):
        return "", ""
    commodity_block = _commodity_block(item)
    if commodity_block:
        title = commodity_block.get("title") or commodity_block.get("theme_name") or ""
        desc = commodity_block.get("description") or commodity_block.get("content") or ""
    else:
        title = item.get("title") or item.get("theme_name") or ""
        desc = item.get("description") or item.get("content") or ""
    return str(title or "").strip(), str(desc or "").strip()


def _source_text_blob(title: str, desc: str) -> str:
    return f"{title}\n{desc}".strip()


def build_ai_source_text(item: dict) -> str:
    """Собирает текст для ИИ: первая ссылка + тексты со всех доп. ссылок."""
    title, desc = extract_item_text(item)
    parts = [f"Заголовок: {title}\nОписание: {desc}"]
    extras = item.get("_extra_sources") if isinstance(item, dict) else None
    if isinstance(extras, list):
        for i, extra in enumerate(extras, 2):
            if isinstance(extra, dict):
                extra_title = str(extra.get("title") or "").strip()
                extra_desc = str(extra.get("description") or "").strip()
            else:
                extra_title, extra_desc = "", str(extra or "").strip()
            if not extra_title and not extra_desc:
                continue
            parts.append(
                f"--- Источник #{i} ---\nЗаголовок: {extra_title}\nОписание: {extra_desc}"
            )
    return "\n\n".join(parts)


_EXAMPLE_LEAK_RE = re.compile(
    r"speedy\s*p9|escale\s*antigua|bandouli[eè]re\s*25|"
    r"ультрасовременное переосмысление культового силуэта|"
    r"vcard31800|alhambra|"
    r"четырехлистного клевера",
    re.IGNORECASE,
)
_CLOTHING_SRC_RE = re.compile(
    r"t恤|tee\b|t-shirt|短袖|卫衣|外套|夹克|羽绒服|hoodie|jacket|coat|"
    r"курт|футболк|худи|свитш|поло\b|рубашк|пальто|жилет",
    re.IGNORECASE,
)
_BAG_OUT_RE = re.compile(r"сумк|speedy|\bbag\b|тоут|clutch|handbag", re.IGNORECASE)


def meaningful_source_payload(raw_text: str) -> str:
    """Текст поставщика без служебных подписей «Заголовок/Описание/Источник»."""
    text = str(raw_text or "")
    text = re.sub(r"--- Источник #\d+ ---", " ", text)
    text = re.sub(r"Заголовок:\s*", " ", text)
    text = re.sub(r"Описание:\s*", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def source_text_is_thin(raw_text: str) -> bool:
    """Мало/нет текста поставщика — ИИ начнёт копировать примеры из промпта."""
    return len(meaningful_source_payload(raw_text)) < 12


def _card_text_blob(card_data: dict) -> str:
    return " ".join(
        str(card_data.get(key) or "")
        for key in ("title_ru_short", "title_ru", "description_ru", "brand")
    )


def sanitize_hallucinated_card(card_data: dict, source_text: str) -> bool:
    """
    Если ИИ скопировал пример из промпта (сумка LV / Alhambra) или написал сумку
    по одежде — обнуляет выдуманные текстовые поля. True, если почистили.
    """
    source = source_text or ""
    blob = _card_text_blob(card_data)
    leaked = bool(_EXAMPLE_LEAK_RE.search(blob)) and not _EXAMPLE_LEAK_RE.search(source)
    type_mismatch = bool(_CLOTHING_SRC_RE.search(source)) and bool(_BAG_OUT_RE.search(blob))
    if not leaked and not type_mismatch:
        return False

    card_data["description_ru"] = ""
    if leaked or _BAG_OUT_RE.search(str(card_data.get("title_ru_short") or "") + " " + str(card_data.get("title_ru") or "")):
        card_data["title_ru_short"] = ""
        card_data["title_ru"] = ""
    brand = str(card_data.get("brand") or "").strip()
    if leaked and brand.lower() in {"louis vuitton", "van cleef & arpels"} and not re.search(
        r"louis\s*vuitton|\blv\b|van\s*cleef|vca", source, re.IGNORECASE
    ):
        card_data["brand"] = ""
    tags = card_data.get("tags")
    if isinstance(tags, list):
        card_data["tags"] = [t for t in tags if not _EXAMPLE_LEAK_RE.search(str(t)) and not _BAG_OUT_RE.search(str(t))]
    elif leaked:
        card_data["tags"] = []
    return True


def extract_image_urls(item: dict) -> List[str]:
    """Достаёт URL фото из сырого товара (commodity или плоская структура)."""
    if not isinstance(item, dict):
        return []

    img_urls = []
    commodity_data = _commodity_block(item)
    if commodity_data:
        img_urls = commodity_data.get("imgsSrc") or commodity_data.get("imgs") or []

    if not img_urls:
        img_urls = item.get("imgsSrc") or item.get("imgs") or item.get("image_list") or item.get("img_list") or []

    return _normalize_img_list(img_urls)


def set_image_urls(item: dict, urls: List[str]) -> None:
    """Записывает список фото в сырой товар (commodity.imgsSrc или плоский imgsSrc)."""
    if "commodity" in item and isinstance(item["commodity"], dict):
        item["commodity"]["imgsSrc"] = list(urls)
        return
    result = item.get("result") if isinstance(item.get("result"), dict) else None
    if result and isinstance(result.get("commodity"), dict):
        result["commodity"]["imgsSrc"] = list(urls)
        return
    item["imgsSrc"] = list(urls)


def resolve_item_id(item: dict, idx: int = 1) -> str:
    """Единый способ достать szwego/goods id из сырой записи."""
    commodity_block = {}
    if isinstance(item, dict):
        commodity_block = item.get("commodity") or {}
        if not commodity_block:
            result = item.get("result") if isinstance(item.get("result"), dict) else {}
            commodity_block = result.get("commodity") if isinstance(result.get("commodity"), dict) else {}
    return str(
        item.get("goods_id")
        or item.get("id")
        or item.get("itemId")
        or (commodity_block.get("goods_id") if isinstance(commodity_block, dict) else None)
        or (commodity_block.get("id") if isinstance(commodity_block, dict) else None)
        or f"unknown_{idx}"
    )


def save_to_raw_dump(item_data: dict, original_url: str):
    """Сохраняет сырой товар в черновик json-файла (накопительный режим)"""
    data = load_raw_dump()

    # Защита: вытаскиваем уникальный ID товара из самой ссылки Павла, если API его не вернуло
    item_id_match = re.search(r"(?:theme_detail|goods_detail)/[^/]+/([^/?#]+)", original_url)
    fallback_id = item_id_match.group(1) if item_id_match else f"id_{int(time.time()*1000)}"

    current_id = item_data.get("id") or item_data.get("itemId") or fallback_id
    item_data["id"] = current_id  # Гарантируем наличие ключа id для всей системы

    if any(str(x.get("id")) == str(current_id) for x in data):
        print("ℹ️ Этот товар уже есть в списке черновиков.")
        return

    data.append(item_data)
    save_raw_dump(data)
    print(f"✅ Товар добавлен в черновик (Всего в списке: {len(data)}).")


async def run_single_mode(url: str):
    """Точка входа для поштучного режима"""
    if not url.startswith("http"):
        print("❌ Строка не похожа на ссылку! Проверьте ввод.")
        return

    item_data = await fetch_single_item_raw_url(url)
    if item_data:
        # Передаем url аргументом для генерации fallback_id
        save_to_raw_dump(item_data, url)
    else:
        print("❌ Не удалось получить данные товара.")


def _merge_extra_source_text(target: dict, new_item: dict, url: str) -> bool:
    """Дописывает текст доп. ссылки в _extra_sources, если он новый."""
    incoming_title, incoming_desc = extract_item_text(new_item)
    blob = _source_text_blob(incoming_title, incoming_desc)
    if not blob:
        return False

    first_title, first_desc = extract_item_text(target)
    seen = {_source_text_blob(first_title, first_desc)}
    extras = list(target.get("_extra_sources") or [])
    for extra in extras:
        if isinstance(extra, dict):
            seen.add(_source_text_blob(str(extra.get("title") or ""), str(extra.get("description") or "")))
        else:
            seen.add(str(extra or "").strip())

    if blob in seen:
        return False

    extras.append({"title": incoming_title, "description": incoming_desc, "url": url})
    target["_extra_sources"] = extras
    return True


async def run_append_link_mode():
    """П.3: докинуть фото и текст с другой ссылки в последний товар черновика."""
    data = load_raw_dump()
    if not data:
        print("❌ Черновик пуст. Сначала добавьте товар пунктом 1.")
        return

    url = input("👉 Вставьте ссылку с доп. фото/описанием для ПОСЛЕДНЕГО товара: ").strip()
    if not url.startswith("http"):
        print("❌ Строка не похожа на ссылку! Проверьте ввод.")
        return

    new_item = await fetch_single_item_raw_url(url)
    if not new_item:
        print("❌ Не удалось получить данные по ссылке.")
        return

    target = data[-1]
    existing = extract_image_urls(target)
    incoming = extract_image_urls(new_item)

    seen = set(existing)
    added = []
    for img_url in incoming:
        if img_url and img_url not in seen:
            seen.add(img_url)
            added.append(img_url)

    added_text = _merge_extra_source_text(target, new_item, url)

    if not added and not added_text:
        print("ℹ️ Новых фото и текста не найдено (всё уже есть в последнем товаре).")
        return

    if added:
        merged = existing + added
        set_image_urls(target, merged)
    else:
        merged = existing

    save_raw_dump(data)
    extra_count = len(target.get("_extra_sources") or [])
    print(
        f"✅ К последнему товару добавлено фото: {len(added)}. "
        f"Всего фото: {len(merged)}. "
        f"Доп. текстов для ИИ: {extra_count}. "
        f"Товаров в черновике: {len(data)}."
    )
    if added_text:
        print("   Текст с этой ссылки тоже пойдёт в описание при сборке таблицы.")


async def run_set_manual_meta_mode():
    """П.4: вручную задать название, бренд и категорию для последнего товара."""
    data = load_raw_dump()
    if not data:
        print("❌ Черновик пуст. Сначала добавьте товар пунктом 1.")
        return

    title = input("👉 Короткое название товара: ").strip()
    brand = input("👉 Бренд: ").strip()
    if not title or not brand:
        print("❌ Нужны и название, и бренд. Повторите пункт 4.")
        return

    target = data[-1]
    target["_manual_export"] = True
    target["_manual_title_ru_short"] = title
    target["_manual_brand"] = brand
    target["_manual_category"] = prompt_category_path()
    save_raw_dump(data)
    print(f"✅ Для последнего товара записано вручную: «{title}» / {brand}")
    print(f"   Категория: {target['_manual_category']}")
    print("   При сборке таблицы название, бренд и категория уйдут как введено.")

# Разрешённые пути категорий строго по дереву из ai_prompt.txt
# Разделитель иерархии категорий для сайта: строго ">" без пробелов вокруг.
# Пробелы внутри названия сегмента (напр. "Футболки, майки, топы") сохраняются.
CATEGORY_SEP = ">"

_CATEGORY_LEAVES = [
    "Одежда>Верхняя одежда>Ветровки",
    "Одежда>Верхняя одежда>Пуховики",
    "Одежда>Верхняя одежда>Пальто",
    "Одежда>Верхняя одежда>Кожаные куртки",
    "Одежда>Верхняя одежда>Джуты",
    "Одежда>Верхняя одежда>Жилеты",
    "Одежда>Верхняя одежда>Шубы",
    "Одежда>Худи>На замке",
    "Одежда>Худи>Без замка",
    "Одежда>Свитшоты",
    "Одежда>Брюки>Повседневные",
    "Одежда>Брюки>Классические",
    "Одежда>Брюки>Спортивные брюки",
    "Одежда>Шорты",
    "Одежда>Плавки и купальники>Плавки",
    "Одежда>Плавки и купальники>Купальники",
    "Одежда>Футболки, майки, топы>Футболки",
    "Одежда>Футболки, майки, топы>Поло",
    "Одежда>Футболки, майки, топы>Майки",
    "Одежда>Футболки, майки, топы>Топы",
    "Одежда>Лонгсливы",
    "Одежда>Рубашки и блузки>Длинный рукав",
    "Одежда>Рубашки и блузки>Короткий рукав",
    "Одежда>Свитеры",
    "Одежда>Джинсы",
    "Одежда>Платья",
    "Одежда>Юбки",
    "Обувь>Кроссовки",
    "Обувь>Кеды",
    "Обувь>Лоферы",
    "Обувь>Мюли",
    "Обувь>Ботинки",
    "Обувь>Мокасины",
    "Обувь>Сандалии",
    "Обувь>Туфли",
    "Обувь>Шлепанцы",
    "Обувь>Сапоги",
    "Обувь>Балетки",
    "Обувь>Ботильоны",
    "Обувь>Босоножки",
    "Аксессуары>Сумки и рюкзаки>Через плечо",
    "Аксессуары>Сумки и рюкзаки>Поясные",
    "Аксессуары>Сумки и рюкзаки>Барсетки",
    "Аксессуары>Сумки и рюкзаки>Мессенджеры",
    "Аксессуары>Сумки и рюкзаки>Дорожные и чемоданы",
    "Аксессуары>Сумки и рюкзаки>Рюкзаки",
    "Аксессуары>Сумки и рюкзаки>Клатчи",
    "Аксессуары>Сумки и рюкзаки>Косметички",
    "Аксессуары>Кошельки>Кошельки",
    "Аксессуары>Кошельки>Картхолдеры",
    "Аксессуары>Кошельки>Визитницы",
    "Аксессуары>Часы",
    "Аксессуары>Ремни",
    "Аксессуары>Очки",
    "Аксессуары>Головные уборы>Кепки",
    "Аксессуары>Головные уборы>Панамы",
    "Аксессуары>Головные уборы>Шапки",
    "Аксессуары>Головные уборы>Шляпы",
]
_CATEGORY_GENDERS = ("Мужской", "Женский", "Унисекс")
ALLOWED_CATEGORIES = {
    f"{gender}{CATEGORY_SEP}{leaf}"
    for gender in _CATEGORY_GENDERS
    for leaf in _CATEGORY_LEAVES
}

# Короткие/устаревшие названия → ближайший путь из дерева (пол уточняется отдельно)
_CATEGORY_ALIASES = {
    "сумки": "Аксессуары>Сумки и рюкзаки>Через плечо",
    "сумка": "Аксессуары>Сумки и рюкзаки>Через плечо",
    "рюкзаки": "Аксессуары>Сумки и рюкзаки>Рюкзаки",
    "кошельки": "Аксессуары>Кошельки>Кошельки",
    "портмоне": "Аксессуары>Кошельки>Кошельки",
    "карты": "Аксессуары>Кошельки>Картхолдеры",
    "картхолдер": "Аксессуары>Кошельки>Картхолдеры",
    "часы": "Аксессуары>Часы",
    "ремни": "Аксессуары>Ремни",
    "очки": "Аксессуары>Очки",
    "обувь": "Обувь>Кеды",
    "кроссовки": "Обувь>Кроссовки",
    "кеды": "Обувь>Кеды",
    "лоферы": "Обувь>Лоферы",
    "мюли": "Обувь>Мюли",
    "ботинки": "Обувь>Ботинки",
    "туфли": "Обувь>Туфли",
    "шлепанцы": "Обувь>Шлепанцы",
    "сандалии": "Обувь>Сандалии",
    "одежда и аксессуары": "Одежда>Футболки, майки, топы>Футболки",
    "худи": "Одежда>Худи>Без замка",
    "свитшот": "Одежда>Свитшоты",
    "свитер": "Одежда>Свитеры",
    "кардиган": "Одежда>Свитеры",
    "джемпер": "Одежда>Свитеры",
    "водолазка": "Одежда>Свитеры",
    "футболка": "Одежда>Футболки, майки, топы>Футболки",
    "поло": "Одежда>Футболки, майки, топы>Поло",
    "джинсы": "Одежда>Джинсы",
    "платье": "Одежда>Платья",
    "юбка": "Одежда>Юбки",
}


def _category_tree() -> dict:
    """Дерево выбора из _CATEGORY_LEAVES: узел — dict, лист — None."""
    tree: dict = {}
    for leaf in _CATEGORY_LEAVES:
        parts = [p for p in leaf.split(CATEGORY_SEP) if p]
        node = tree
        for i, part in enumerate(parts):
            is_last = i == len(parts) - 1
            if is_last:
                node.setdefault(part, None)
            else:
                child = node.get(part)
                if not isinstance(child, dict):
                    node[part] = {}
                node = node[part]
    return tree


def _prompt_menu_choice(title: str, options: List[str], current_path: str = "") -> int:
    """Нумерованное меню. Возвращает индекс выбранного пункта."""
    while True:
        print()
        if current_path:
            print(f"Сейчас: {current_path}")
        print(title)
        for i, option in enumerate(options, 1):
            print(f"{i}. {option}")
        raw = input(f"👉 Номер (1-{len(options)}): ").strip()
        if raw.isdigit() and 1 <= int(raw) <= len(options):
            return int(raw) - 1
        print("❌ Неверный номер, попробуйте ещё раз.")


def prompt_category_path() -> str:
    """Интерактивный выбор категории из дерева (пол + путь)."""
    print("\n📂 Категория")
    gender = _CATEGORY_GENDERS[_prompt_menu_choice("Пол:", list(_CATEGORY_GENDERS))]

    node = _category_tree()
    chosen: List[str] = []
    while True:
        keys = list(node.keys())
        labels = [
            key if node[key] is None else f"{key} ..."
            for key in keys
        ]
        current = _join_category_parts([gender] + chosen)
        key = keys[_prompt_menu_choice("Раздел:", labels, current)]
        chosen.append(key)
        child = node[key]
        if child is None:
            path = _join_category_parts([gender] + chosen)
            print(f"✅ Категория: {path}")
            return path
        node = child


def _split_category_parts(value: str) -> List[str]:
    """Режет путь по '>' / '/' / стрелкам; пробелы вокруг разделителя игнорирует."""
    cleaned = re.sub(r"\s*[>/→\-–—]+\s*", CATEGORY_SEP, str(value or "").strip())
    cleaned = re.sub(r"\s+", " ", cleaned).strip(CATEGORY_SEP + " ")
    return [p.strip() for p in cleaned.split(CATEGORY_SEP) if p.strip()]


def _join_category_parts(parts: List[str]) -> str:
    """Склеивает сегменты как Унисекс>Одежда>Футболки (без пробелов вокруг '>')."""
    return CATEGORY_SEP.join(p.strip() for p in parts if p and str(p).strip())


def normalize_category_path(value: str) -> str:
    """Нормализует путь категории к дереву из промпта. Разделитель: '>' без пробелов."""
    raw = str(value or "").strip()
    if not raw:
        return ""

    parts = _split_category_parts(raw)
    cleaned_for_gender = " ".join(parts)

    gender = "Унисекс"
    leaf_parts = parts
    if parts and parts[0] in _CATEGORY_GENDERS:
        gender = parts[0]
        leaf_parts = parts[1:]
    elif "женщин" in cleaned_for_gender.lower() or "женск" in cleaned_for_gender.lower():
        gender = "Женский"
    elif "мужчин" in cleaned_for_gender.lower() or "мужск" in cleaned_for_gender.lower():
        gender = "Мужской"

    leaf = _join_category_parts(leaf_parts)
    candidate = _join_category_parts([gender, leaf]) if leaf else ""
    if candidate in ALLOWED_CATEGORIES:
        return candidate

    # Точное совпадение листа без пола
    for allowed_leaf in _CATEGORY_LEAVES:
        if leaf.lower() == allowed_leaf.lower():
            return _join_category_parts([gender, allowed_leaf])

    # Алиасы коротких названий
    compact = leaf.lower() if leaf else cleaned_for_gender.lower()
    for alias, mapped_leaf in _CATEGORY_ALIASES.items():
        if alias in compact:
            return _join_category_parts([gender, mapped_leaf])

    # Частичное совпадение по последнему сегменту (Кроссовки, Через плечо, ...)
    last = (leaf_parts[-1] if leaf_parts else cleaned_for_gender).lower()
    for allowed_leaf in _CATEGORY_LEAVES:
        if allowed_leaf.split(CATEGORY_SEP)[-1].lower() == last:
            return _join_category_parts([gender, allowed_leaf])
        if last in allowed_leaf.lower():
            return _join_category_parts([gender, allowed_leaf])

    return ""


def format_card_for_export(card_data: dict) -> dict:
    """Подготавливает строку для экспорта: пустой sku, теги через запятую, абзацы в description сохраняются."""
    row = dict(card_data)
    sku = str(row.get("sku", "")).strip()
    if not sku or sku.upper() == "N/A":
        row["sku"] = ""
    tags = row.get("tags", [])
    if isinstance(tags, list):
        row["tags"] = ", ".join(str(tag).strip() for tag in tags if str(tag).strip())
    elif tags is None:
        row["tags"] = ""
    else:
        row["tags"] = str(tags).strip()

    normalized = normalize_category_path(str(row.get("category", "")))
    if normalized:
        row["category"] = normalized
    return row


def export_products_table(export_rows: list) -> str:
    """
    Пишет xlsx: абзацы (\\n) внутри ячейки description сохраняются,
    но Wrap Text выключен — строки не растягиваются при просмотре в Excel.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment

    if not export_rows:
        raise ValueError("Нет строк для экспорта")

    columns = list(export_rows[0].keys())
    wb = Workbook()
    ws = wb.active
    ws.title = "products"

    # Заголовки
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.alignment = Alignment(wrap_text=False, vertical="center")

    # Данные
    for row_idx, row in enumerate(export_rows, 2):
        for col_idx, header in enumerate(columns, 1):
            value = row.get(header, "")
            if value is None:
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Переносы внутри значения остаются, визуальный wrap в Excel — выключен
            cell.alignment = Alignment(wrap_text=False, vertical="center")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_name = f"products_export_{timestamp}.xlsx"
    wb.save(excel_name)
    return excel_name

async def process_and_export_table(raw_items: list):
    """Принимает список сырых товаров, прогоняет через ИИ (или manual) и делает экспорт"""
    if not raw_items:
        print("❌ Нет данных для обработки!")
        return

    print(f"🤖 Запуск пакетной обработки для {len(raw_items)} товаров...")
    final_cards = []
    failed_items = []
    failed_raw_items = []

    for idx, item in enumerate(raw_items, 1):
        title, _desc = extract_item_text(item) if isinstance(item, dict) else ("", "")

        item_id = resolve_item_id(item, idx)
        img_urls = extract_image_urls(item)
        original_imgs = ", ".join(img_urls) if img_urls else ""
        manual_category = str(item.get("_manual_category") or "").strip() if isinstance(item, dict) else ""

        # Ручной режим: название/бренд от заказчика, без ИИ
        if item.get("_manual_export"):
            manual_title = str(item.get("_manual_title_ru_short") or "").strip()
            manual_brand = str(item.get("_manual_brand") or "").strip()
            print(f" ✍️ [{idx}/{len(raw_items)}] Ручной товар ID: {item_id}...")
            if not manual_title or not manual_brand:
                failed_items.append({
                    "szwego_id": str(item_id),
                    "title": str(title)[:200],
                    "reason": "Ручной режим: не заданы название и/или бренд",
                })
                failed_raw_items.append(item)
                print(f"⚠️ [SKIP] Ручной товар без названия/бренда ID: {item_id}")
                continue

            card_data = {
                "title_ru_short": manual_title,
                "title_ru": manual_title,
                "brand": manual_brand,
                "sku": "",
                "description_ru": "",
                "category": manual_category,
                "tags": [],
                "szwego_id": item_id,
                "original_imgs": original_imgs,
            }
            final_cards.append(card_data)
            print(f" -> [OK MANUAL] {manual_title} / {manual_brand} / {manual_category or 'без категории'}")
            continue

        # 3. Собираем финальный текст для отправки в OpenRouter (все ссылки товара)
        raw_text = build_ai_source_text(item)

        if source_text_is_thin(raw_text):
            print(f" ℹ️ [{idx}/{len(raw_items)}] Нет текста поставщика ID: {item_id} — описание не выдумываю.")
            card_data = {
                "title_ru_short": str(title)[:200] if title else "",
                "title_ru": str(title) if title else "",
                "brand": "",
                "sku": "",
                "description_ru": "",
                "category": manual_category,
                "tags": [],
                "szwego_id": item_id,
                "original_imgs": original_imgs,
            }
            final_cards.append(card_data)
            print(" -> [OK EMPTY] Описание пустое (нет исходного текста)")
            continue

        print(f" 🤖 [{idx}/{len(raw_items)}] Обработка товара ID: {item_id}...")
        ai_card = await generate_ai_card(raw_text)

        if ai_card:
            card_data = ai_card.model_dump() if hasattr(ai_card, 'model_dump') else ai_card
            card_data['szwego_id'] = item_id
            card_data['original_imgs'] = original_imgs
            if manual_category:
                card_data['category'] = manual_category
            if sanitize_hallucinated_card(card_data, raw_text):
                print("    ⚠️ Похоже на выдумку/пример из промпта — текстовые поля очищены.")
            final_cards.append(card_data)
            print(f" -> [OK] {card_data.get('title_ru', '')} (SKU: {card_data.get('sku', '')})")
        else:
            failed_items.append({
                "szwego_id": str(item_id),
                "title": str(title)[:200],
                "reason": "ИИ не вернул валидный JSON после 3 попыток",
            })
            failed_raw_items.append(item)
            print(f"⚠️ [SKIP] Ошибка генерации для товара ID: {item_id}")

    if failed_items:
        fail_stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        failed_report = f"failed_products_{fail_stamp}.json"
        failed_raw_file = f"failed_raw_products_{fail_stamp}.json"
        with open(failed_report, "w", encoding="utf-8") as f:
            json.dump(failed_items, f, ensure_ascii=False, indent=2)
        with open(failed_raw_file, "w", encoding="utf-8") as f:
            json.dump(failed_raw_items, f, ensure_ascii=False, indent=2)
        print(f"\n⚠️ Не сгенерировано товаров: {len(failed_items)} из {len(raw_items)}")
        print(f"   Список ID/названий: {failed_report}")
        print(f"   Сырые данные для повтора: {failed_raw_file}")
        print("   Чтобы перегенерировать только их: скопируйте файл в raw_selected_products.json и снова пункт 4.")

    if not final_cards:
        print("❌ Ни один товар не был успешно обработан.")
        return

    with open("final_products.json", "w", encoding="utf-8") as f:
        json.dump(final_cards, f, ensure_ascii=False, indent=4)

    export_rows = [format_card_for_export(card) for card in final_cards]
    excel_name = export_products_table(export_rows)
    print(f"\n📊 БОМБА! Итоговая таблица создана: {excel_name}")
    print(f"   Успешно: {len(final_cards)} | Ошибки: {len(failed_items)}")

async def main_cli():
    parser = argparse.ArgumentParser(description="Szwego Pipeline CLI")
    parser.add_argument("--all", action="store_true", help="Парсить весь каталог")
    parser.add_argument("--limit", type=int, help="Переопределить лимит количества товаров")
    parser.add_argument("--single", action="store_true", help="Парсить один товар по ссылке")
    parser.add_argument("--append-link", action="store_true", help="Докинуть фото и текст с другой ссылки в последний товар черновика")
    parser.add_argument("--set-manual-meta", action="store_true", help="Вручную задать название, бренд и категорию для последнего товара")
    parser.add_argument("--build-table", action="store_true", help="Собрать накопленные товары в Excel через ИИ")

    args = parser.parse_args()

    # РЕЖИМ 1: Парсинг всего каталога
    if args.all:
        # Если передан лимит из батника — берем его, иначе — из .env, иначе — 100 по умолчанию
        target_limit = args.limit if args.limit is not None else int(os.getenv("TOTAL_TARGET_PRODUCTS", 100))
        print(f"🚀 Запуск полного парсинга. Цель: {target_limit} товаров.")

        # --- ЦИКЛ ПАГИНАЦИИ (STAGE-1) ---
        await main(target_limit)

    # РЕЖИМ 2: Поштучный сбор по ссылке
    elif args.single:
        url_arg = input("👉 Вставьте ссылку на товар Szwego: ").strip()
        if url_arg:
            await run_single_mode(url_arg)

    # РЕЖИМ 2b: доп. ссылка к последнему товару (фото + текст)
    elif args.append_link:
        await run_append_link_mode()

    # РЕЖИМ 2c: ручные название/бренд/категория к последнему товару
    elif args.set_manual_meta:
        await run_set_manual_meta_mode()

    # РЕЖИМ 3: Пакетная сборка таблицы из поштучных черновиков
    elif args.build_table:
        raw_selected_items = load_raw_dump()
        if not raw_selected_items:
            print(f"❌ Черновик {RAW_DUMP_FILE} не найден или пуст! Сначала добавьте товары через пункт 3.")
            return

        print(f"📂 Загружено {len(raw_selected_items)} товаров из поштучного списка черновиков.")
        await process_and_export_table(raw_selected_items)

        # По желанию: очищаем файл-черновик после успешного экспорта в Excel
        try:
            os.remove(RAW_DUMP_FILE)
            print(f"🧹 Черновик {RAW_DUMP_FILE} успешно очищен.")
        except Exception:
            pass
    else:
        parser.print_help()

if __name__ == "__main__":
    # Заменяем старый запуск asyncio.run(main()) на новый CLI-обработчик
    asyncio.run(main_cli())